# Automações em Python

'''
Vantagens de usar python para automações:

- Facilidade de aprendizado
- Linguagem de auto nível (ou seja, fácil de entender)
- Comunidade ativa
- Bibliotecas prontas para automações
- Multiplataforma
- Gratuito
- Open Source
- Fácil de instalar
- Fácil de usar
'''

# Essa biblioteca é utilizada para manipular arquivos do Excel
from openpyxl import load_workbook

# Caminho para o arquivo do excel
arquivo = 'C:/Users/Aluno/Desktop/cursoPython/exelCursoPython.xlsx'

# Carregar o arquivo do excel
wb = load_workbook(arquivo)

# Selecionar a planilha ativa (a primeira planilha por padrão)
planilha = wb.active 

# Iterar sobre as células da planilha e imprimir seus valores

print('Forma realizada com IA:')

for linha in planilha.iter_rows(values_only=True):
    print(linha)

print('\n')    
    
'''
Explicação da lógica e dos simbolos utilizados no código abaixo:

for linha in planilha.iter_rows(values_only=True):
    print(linha)
    
Enquanto houver linhas na planilha, faça:
    Imprima a linha

O .iter_rows() é um método que itera sobre as linhas da planilha
O values_only=True é um argumento que retorna apenas os valores das células
'''

print('Foma do Professor:')

dados = []

# Iterar sobre todas as linhas e colunas
for row in planilha.iter_rows():
    dados.append([cell.value for cell in row])
    
print(dados)

print('\n')    

print('Segunda forma do Professor:')

# Exibir as primeiras linhas de dados
for linha in dados:
    print(linha)
    
print('\n')  

# Com pandas

print('Forma com Pandas:')

import pandas as pd

dados = []

# Corrigir o caminho do arquivo para ser uma string
tabela = pd.read_excel('C:/Users/Aluno/Desktop/cursoPython/exelCursoPython.xlsx')

print(tabela)

print('\n')    