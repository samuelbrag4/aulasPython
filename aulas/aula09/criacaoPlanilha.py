from openpyxl import Workbook, load_workbook

# Cria um novo arquivo do excel (planilha)
wb = Workbook()

# Selecionar a planilha ativa (a primeira planilha por padrão)
planilha = wb.active
planilha.title = 'planilha_feita_com_python'

# Adiciona dados a planilha
planilha['A1'] = 'Nome'
planilha['B1'] = 'Idade'
planilha['C1'] = 'Sexo'

# Adiciona mais dados a planilha
planilha['A2'] = 'João'
planilha['B2'] = 25
planilha['C2'] = 'M'

planilha['A3'] = 'Maria'
planilha['B3'] = 30
planilha['C3'] = 'F'

planilha['A4'] = 'José'
planilha['B4'] = 22
planilha['C4'] = 'M'

planilha['A5'] = 'Ana'
planilha['B5'] = 35
planilha['C5'] = 'F'

# Salva o arquivo
wb.save('C:/Users/Aluno/Desktop/cursoPython/exelCursoPython.xlsx')

# Carregar o arquivo do excel
wb = load_workbook('C:/Users/Aluno/Desktop/cursoPython/exelCursoPython.xlsx')

# Selecionar a planilha ativa (a primeira planilha por padrão)
planilha = wb.active

# Iterar sobre as células da planilha e imprimir seus valores
for linha in planilha.iter_rows(values_only=True):
    print(linha)
    
print('\n')

# Com pandas 
import pandas as pd

df = pd.read_excel('C:/Users/Aluno/Desktop/cursoPython/exelCursoPython.xlsx')

print(df)

print('\n')

# Outra forma mais complexa

dados = {
    'Nome': ['João', 'Maria', 'José', 'Jane'],
    'Idade': [25, 30, 22, 35],
    'Sexo': ['M', 'F', 'M', 'F']
}

df = pd.DataFrame(dados)

# Salvar o DataFrame em um arquivo do excel
df.to_excel('C:/Users/Aluno/Desktop/cursoPython/planilha_feita_com_pandas.xlsx', index=False, sheet_name='planilha_feita_com_pandas')

print('Planilha criada com sucesso!')